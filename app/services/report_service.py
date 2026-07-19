"""
Reporting service: builds CSV and Excel (.xlsx) exports (in-memory,
streamed to the client) and PDF summary reports (via ReportLab) for
inventory and transaction data.
"""
import csv
import io
import logging
from datetime import datetime
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from app.models.product import Product
from app.models.transaction import Transaction

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header_row(sheet, num_columns: int) -> None:
    for col in range(1, num_columns + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize_columns(sheet, num_columns: int, min_width: int = 10, max_width: int = 45) -> None:
    for col in range(1, num_columns + 1):
        letter = get_column_letter(col)
        longest = max(
            (len(str(cell.value)) for cell in sheet[letter] if cell.value is not None),
            default=min_width,
        )
        sheet.column_dimensions[letter].width = max(min_width, min(longest + 2, max_width))


class ReportService:
    @staticmethod
    def products_to_csv(products: List[Product]) -> io.StringIO:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(
            ["Product Code", "Name", "Category", "Price", "Reorder Level",
             "Total Quantity", "Total Value", "Low Stock", "Active"]
        )
        for p in products:
            writer.writerow([
                p.product_code, p.name, p.category.name if p.category else "",
                f"{p.price:.2f}", p.reorder_level, p.total_quantity,
                f"{p.total_value:.2f}", "YES" if p.is_low_stock else "NO",
                "YES" if p.is_active else "NO",
            ])
        buffer.seek(0)
        return buffer

    @staticmethod
    def transactions_to_csv(transactions: List[Transaction]) -> io.StringIO:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(
            ["Date", "Product Code", "Type", "Quantity", "Before", "After",
             "Unit Price", "Reference", "Performed By (user id)", "Notes"]
        )
        for t in transactions:
            writer.writerow([
                t.created_at.strftime("%Y-%m-%d %H:%M"),
                t.product.product_code if t.product else "",
                t.transaction_type.value,
                t.quantity, t.quantity_before, t.quantity_after,
                f"{t.unit_price_snapshot:.2f}", t.reference or "",
                t.performed_by, t.notes or "",
            ])
        buffer.seek(0)
        return buffer

    @staticmethod
    def products_to_excel(products: List[Product]) -> io.BytesIO:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Products"

        headers = ["Product Code", "Name", "Category", "Price", "Reorder Level",
                   "Total Quantity", "Total Value", "Low Stock", "Active"]
        sheet.append(headers)

        for p in products:
            sheet.append([
                p.product_code, p.name, p.category.name if p.category else "",
                round(p.price, 2), p.reorder_level, p.total_quantity,
                round(p.total_value, 2), "YES" if p.is_low_stock else "NO",
                "YES" if p.is_active else "NO",
            ])

        _style_header_row(sheet, len(headers))
        _autosize_columns(sheet, len(headers))
        sheet.freeze_panes = "A2"  # keep header visible while scrolling

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        logger.info("Generated Excel product export for %s products", len(products))
        return buffer

    @staticmethod
    def transactions_to_excel(transactions: List[Transaction]) -> io.BytesIO:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Transactions"

        headers = ["Date", "Product Code", "Type", "Quantity", "Before", "After",
                   "Unit Price", "Reference", "Performed By (user id)", "Notes"]
        sheet.append(headers)

        for t in transactions:
            sheet.append([
                t.created_at.strftime("%Y-%m-%d %H:%M"),
                t.product.product_code if t.product else "",
                t.transaction_type.value.replace("_", " ").title(),
                t.quantity, t.quantity_before, t.quantity_after,
                round(t.unit_price_snapshot, 2), t.reference or "",
                t.performed_by, t.notes or "",
            ])

        _style_header_row(sheet, len(headers))
        _autosize_columns(sheet, len(headers))
        sheet.freeze_panes = "A2"

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        logger.info("Generated Excel transaction export for %s rows", len(transactions))
        return buffer

    @staticmethod
    def build_inventory_summary_pdf(
        products: List[Product], total_value: float, low_stock_count: int
    ) -> io.BytesIO:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("StockFlow - Inventory Summary Report", styles["Title"]))
        elements.append(
            Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"])
        )
        elements.append(Spacer(1, 12))

        summary_data = [
            ["Total Products", str(len(products))],
            ["Total Inventory Value", f"${total_value:,.2f}"],
            ["Low Stock Items", str(low_stock_count)],
        ]
        summary_table = Table(summary_data, colWidths=[8 * cm, 6 * cm])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f1f5f9")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("PADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        elements.append(Paragraph("Product Detail", styles["Heading2"]))
        table_data = [["Code", "Name", "Qty", "Price", "Value", "Low Stock"]]
        for p in products:
            table_data.append([
                p.product_code, p.name[:30], str(p.total_quantity),
                f"${p.price:,.2f}", f"${p.total_value:,.2f}",
                "YES" if p.is_low_stock else "",
            ])

        detail_table = Table(table_data, colWidths=[2.5 * cm, 6 * cm, 1.8 * cm, 2.5 * cm, 2.8 * cm, 2.5 * cm])
        detail_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("PADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(detail_table)

        doc.build(elements)
        buffer.seek(0)
        logger.info("Generated PDF inventory summary report for %s products", len(products))
        return buffer
